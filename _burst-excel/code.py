# code.py

import sys
from openpyxl import load_workbook

def delete_rows_except_value(file_path, column_value):
    # マスターエクセルファイルの読み込み
    wb_master = load_workbook(file_path)
    ws_master = wb_master.active

    # 条件に基づいて行を削除
    rows_to_delete = []
    for row in ws_master.iter_rows(min_row=1):
        cell_value = row[0].value  # A列の値を取得
        if column_value not in  cell_value:
            rows_to_delete.append(row)

    # 削除する行を逆順にソートしてから削除（インデックスが変わらないようにするため）
    rows_to_delete.sort(reverse=True, key=lambda x: x[0].row)
    for row in rows_to_delete:
        ws_master.delete_rows(row[0].row)

    # 変更を保存
    wb_master.save("updated_master.xlsx")

if __name__ == "__main__":
    # 引数の取得
    if len(sys.argv) != 3:
        print("Usage: python code.py <file_path> <column_value>")
        sys.exit(1)

    file_path = sys.argv[1]
    column_value = sys.argv[2]

    # 関数の呼び出し
    delete_rows_except_value(file_path, column_value)

